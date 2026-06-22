from __future__ import annotations

import csv
import io
from typing import Dict, List, Optional, Tuple

from datetime import date, datetime

from ..models import (
    AnalysisResult,
    Finding,
    PerfSummary,
    PortfolioResult,
    PortfolioSupplier,
    Severity,
    Verdict,
    VerdictKind,
)


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_number(value) -> Optional[float]:
    """Parse messy spreadsheet numerics: '$6,69,93,913', '($84,711)', '0.0302', '1.8E7'."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() in ("nan", "none", "-"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace("%", "").replace(" ", "")
    if s in ("", "-"):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def load_rows(filename: str, content: bytes) -> Tuple[List[Dict[str, str]], List[str]]:
    """Load a CSV or Excel file into (rows-as-dicts, header-list)."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
        headers = list(reader.fieldnames or [])
        return rows, headers

    # Excel
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header_row = next(it, None) or []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    rows: List[Dict[str, str]] = []
    for r in it:
        if r is None:
            continue
        row = {headers[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(headers)}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return rows, headers


# ── Column detection ──────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return "".join(str(s).lower().split())


def _find(headers: List[str], *candidates: str) -> Optional[str]:
    """Find a header that exactly matches one of the candidates (normalised)."""
    norm_map = {_norm(h): h for h in headers}
    for cand in candidates:
        if _norm(cand) in norm_map:
            return norm_map[_norm(cand)]
    return None


def _contains(headers: List[str], *terms: str) -> Optional[str]:
    """Find the first header containing all terms (normalised) and no excluded markers."""
    for h in headers:
        n = _norm(h)
        if all(_norm(t) in n for t in terms):
            return h
    return None


# ── Score-driven verdict (renew / tighten / scrap) ────────────────────────────
#
# Bands (locked with the merchant):
#   score >= 75  -> RENEW   — favorable, long term
#   50 <= score < 75 -> TIGHTEN — salvage with stricter terms + remediation
#   score < 50   -> SCRAP   — recommend exit / non-renewal

RENEW_MIN = 75
TIGHTEN_MIN = 50


def decide_verdict(score: int, findings: List[Finding]) -> Verdict:
    high = sum(1 for f in findings if f.severity == Severity.high)
    issue_summary = (
        f"{len(findings)} performance issue(s) flagged"
        + (f", {high} high-severity" if high else "")
        if findings
        else "no material issues detected"
    )

    if score >= RENEW_MIN:
        return Verdict(
            kind=VerdictKind.renew,
            label="Renew",
            score=score,
            headline="Renew on favorable terms — this supplier is performing well.",
            rationale=(
                f"Health score {score}/100 is at or above the {RENEW_MIN} renewal bar "
                f"({issue_summary}). Reward reliability and lock in continuity."
            ),
            recommended_term="2-year term with auto-renew",
            recommended_actions=[
                "Offer a 2-year renewal with an auto-renew clause.",
                "Keep existing protections; no penalty tightening needed.",
                "Consider preferred-supplier pricing or volume incentives.",
            ],
        )

    if score >= TIGHTEN_MIN:
        return Verdict(
            kind=VerdictKind.tighten,
            label="Tighten",
            score=score,
            headline="Renew, but tighten the terms — salvage with stricter protections.",
            rationale=(
                f"Health score {score}/100 sits in the {TIGHTEN_MIN}–{RENEW_MIN - 1} salvage band "
                f"({issue_summary}). Worth keeping, but only with accountability built in."
            ),
            recommended_term="1-year term with remediation milestones",
            recommended_actions=[
                "Offer a shorter 1-year term with a mid-term performance review.",
                "Apply the confirmed findings below as clause redlines (penalties, return-rate caps, quality obligations).",
                "Set remediation milestones; make renewal contingent on hitting them.",
            ],
        )

    return Verdict(
        kind=VerdictKind.scrap,
        label="Scrap",
        score=score,
        headline="Recommend exit — do not renew this supplier as-is.",
        rationale=(
            f"Health score {score}/100 is below the {TIGHTEN_MIN} viability floor "
            f"({issue_summary}). The performance gap is too wide to fix with clause tweaks alone."
        ),
        recommended_term="Do not auto-renew; serve non-renewal notice",
        recommended_actions=[
            "Serve non-renewal / termination notice within the contractual notice window.",
            "Begin sourcing a replacement supplier before the renewal date.",
            "If you must bridge, renew only month-to-month with strict exit rights and penalties.",
        ],
    )


# ── Portfolio / bulk roster (one row per supplier, with a score) ──────────────

def _days_to_renewal(value) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(s[:10], fmt).date()
            return (d - date.today()).days
        except ValueError:
            continue
    return None


def _norm_date(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return s[:10]


def analyze_portfolio(rows: List[Dict[str, str]], headers: List[str]) -> PortfolioResult:
    """Map a supplier roster (name + performance_score per row) to verdicts."""
    id_col = _find(headers, "supplier_id", "id") or _contains(headers, "supplier", "id")
    name_col = (
        _find(headers, "name", "supplier", "supplier_name", "vendor")
        or _contains(headers, "supplier", "name")
        or _contains(headers, "name")
        or _contains(headers, "supplier")
    )
    email_col = _find(headers, "email") or _contains(headers, "email")
    score_col = (
        _find(headers, "performance_score", "score", "rating", "health", "performance")
        or _contains(headers, "score")
        or _contains(headers, "rating")
    )
    renewal_col = (
        _find(headers, "renewal_date", "renewal", "expiry", "end_date")
        or _contains(headers, "renewal")
        or _contains(headers, "expiry")
    )

    detected = [c for c in [id_col, name_col, email_col, score_col, renewal_col] if c]

    suppliers: List[PortfolioSupplier] = []
    counts = {"renew": 0, "tighten": 0, "scrap": 0}
    for i, r in enumerate(rows):
        raw_score = parse_number(r.get(score_col)) if score_col else None
        if raw_score is None:
            continue
        score = int(max(0, min(100, round(raw_score))))
        name = (str(r.get(name_col)).strip() if name_col else "") or f"Supplier {i + 1}"
        verdict = decide_verdict(score, [])
        counts[verdict.kind.value] += 1
        renewal = _norm_date(r.get(renewal_col)) if renewal_col else None
        suppliers.append(PortfolioSupplier(
            supplier_id=str(r.get(id_col)).strip() if id_col else "",
            name=name,
            email=str(r.get(email_col)).strip() if email_col else "",
            score=score,
            renewal_date=renewal,
            days_to_renewal=_days_to_renewal(renewal),
            verdict=verdict,
        ))

    # Worst first — surface the suppliers that need a decision soonest.
    suppliers.sort(key=lambda s: s.score)
    return PortfolioResult(suppliers=suppliers, counts=counts, columns_detected=detected)


# ── Analysis (returns & quality — data-backed) ────────────────────────────────

def _col_sum(rows, col) -> float:
    if not col:
        return 0.0
    total = 0.0
    for r in rows:
        v = parse_number(r.get(col))
        if v is not None:
            total += v
    return total


def analyze(rows: List[Dict[str, str]], headers: List[str], label: str) -> AnalysisResult:
    item_col = _find(headers, "Item") or _contains(headers, "item") or _contains(headers, "sku")

    sales_cur_col = _find(headers, "Sales $ (Current Period)") or _contains(headers, "sales", "current")
    sales_prev_col = _find(headers, "Sales $ (Previous Period)") or _contains(headers, "sales", "previous")
    ret_cur_col = _find(headers, "Return $ (Current Period)")
    ret_prev_col = _find(headers, "Return $ (Previous Period)")
    ret_pct_cur_col = _find(headers, "Return $ % (Current Period)")
    ret_pct_asrt_col = _find(headers, "Return $ % - ASRT (Current Period)", "Return $ % - ASRT(Current Period)")
    rtm_units_col = _find(headers, "RTM Units (Current Period)") or _contains(headers, "rtm", "units", "current")

    detected = [c for c in [item_col, sales_cur_col, sales_prev_col, ret_cur_col,
                            ret_pct_cur_col, ret_pct_asrt_col, rtm_units_col] if c]

    sales_cur = _col_sum(rows, sales_cur_col)
    sales_prev = _col_sum(rows, sales_prev_col)
    ret_cur = _col_sum(rows, ret_cur_col)
    ret_prev = _col_sum(rows, ret_prev_col)

    return_rate_cur = (ret_cur / sales_cur) if sales_cur else None
    return_rate_prev = (ret_prev / sales_prev) if sales_prev else None

    findings: List[Finding] = []
    score = 100

    # 1) Overall return rate level / trend
    if return_rate_cur is not None:
        sev = (
            Severity.high if return_rate_cur >= 0.06
            else Severity.medium if return_rate_cur >= 0.04
            else Severity.low
        )
        worsening = (
            return_rate_prev is not None and return_rate_cur > return_rate_prev
        )
        if sev != Severity.low or worsening:
            trend = ""
            if return_rate_prev is not None:
                delta = (return_rate_cur - return_rate_prev) * 100
                trend = f" ({'up' if delta >= 0 else 'down'} {abs(delta):.2f} pts vs previous period)"
            findings.append(Finding(
                id="return_rate",
                metric="return_rate",
                title="High customer return rate",
                observed=f"{return_rate_cur*100:.2f}% of sales returned{trend}",
                benchmark="target ≤ 4%",
                severity=sev,
                evidence=f"Returns ₹{ret_cur:,.0f} on sales ₹{sales_cur:,.0f}.",
                maps_to_clauses=["returns_warranties", "quality"],
                suggested_action=(
                    "Add a return-rate cap and shift return shipping plus restocking cost to "
                    "the Supplier when returns exceed the agreed benchmark, and tighten the "
                    "Supplier's quality-inspection obligation."
                ),
            ))
            score -= 25 if sev == Severity.high else 12 if sev == Severity.medium else 6
            if worsening:
                score -= 5

    # 2) Items returning worse than their peer-group benchmark (ASRT)
    if item_col and ret_pct_cur_col and ret_pct_asrt_col:
        offenders = []
        for r in rows:
            item_pct = parse_number(r.get(ret_pct_cur_col))
            peer_pct = parse_number(r.get(ret_pct_asrt_col))
            if item_pct is None or peer_pct is None:
                continue
            if item_pct > peer_pct:
                offenders.append((str(r.get(item_col)), item_pct - peer_pct, item_pct))
        if offenders:
            offenders.sort(key=lambda x: x[1], reverse=True)
            top = offenders[:5]
            top_str = "; ".join(
                f"{name[:40]} ({pct*100:.1f}% vs peer {(pct-gap)*100:.1f}%)"
                for name, gap, pct in top
            )
            sev = Severity.high if len(offenders) >= max(20, len(rows) * 0.15) else Severity.medium
            findings.append(Finding(
                id="peer_benchmark",
                metric="return_vs_peer",
                title="Products returning worse than peer benchmark",
                observed=f"{len(offenders)} of {len(rows)} items exceed their assortment return benchmark",
                benchmark="at/below assortment (ASRT) average",
                severity=sev,
                evidence=f"Worst offenders: {top_str}.",
                maps_to_clauses=["returns_warranties", "quality"],
                suggested_action=(
                    "Hold the Supplier accountable for items whose return rate exceeds the "
                    "peer-group benchmark: require a remediation plan and credit for excess returns."
                ),
            ))
            score -= 15 if sev == Severity.high else 8

    # 3) Sales trend
    if sales_cur and sales_prev:
        var = sales_cur - sales_prev
        pct = var / sales_prev * 100 if sales_prev else 0
        if var < 0:
            sev = Severity.medium if pct <= -10 else Severity.low
            findings.append(Finding(
                id="sales_trend",
                metric="sales_trend",
                title="Declining sales with this supplier",
                observed=f"Sales down {abs(pct):.1f}% period-over-period",
                benchmark="flat or growing",
                severity=sev,
                evidence=f"Sales ₹{sales_cur:,.0f} vs ₹{sales_prev:,.0f} previously.",
                maps_to_clauses=["pricing_payments"],
                suggested_action=(
                    "Renegotiate pricing/fees or shorten the renewal term given declining "
                    "performance, with a review checkpoint."
                ),
            ))
            score -= 10 if sev == Severity.medium else 4

    # 4) Returns-to-merchant volume (defect signal)
    rtm_units = _col_sum(rows, rtm_units_col)
    if rtm_units_col and rtm_units > 0:
        findings.append(Finding(
            id="rtm_volume",
            metric="rtm",
            title="Material returns-to-merchant volume",
            observed=f"{rtm_units:,.0f} RTM units in the current period",
            benchmark="minimise via quality control",
            severity=Severity.low,
            evidence="Returns flowing back to the merchant indicate quality/fulfilment issues.",
            maps_to_clauses=["quality", "damaged_lost_goods"],
            suggested_action=(
                "Strengthen quality-control and inspection obligations and confirm the Supplier "
                "bears the cost of defective or damaged goods."
            ),
        ))
        score -= 4

    score = max(0, min(100, score))

    summary = PerfSummary(
        supplier_label=label,
        rows=len(rows),
        columns_detected=detected,
        sales_current=sales_cur or None,
        sales_previous=sales_prev or None,
        return_rate_current=return_rate_cur,
        return_rate_previous=return_rate_prev,
        rollup_score=score,
    )
    verdict = decide_verdict(score, findings)
    return AnalysisResult(summary=summary, verdict=verdict, findings=findings)
